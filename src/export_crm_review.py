"""Export the Act! -> SEC CRD matches that still need a human, with the evidence.

WHY
---
A previous pass matched all 47,466 Act! contacts against the SEC adviser feed and
sorted the results into three tiers: high (confident), review (a candidate worth
a second opinion), none (no candidate at all). This turns the review tier into a
worksheet somebody can actually work through, by putting the Act! record and the
proposed SEC record side by side with the reasons the match was uncertain.

The point is to make each row decidable in seconds. A reviewer should not have to
open two systems to answer "is this the same person" -- the firm, the city, the
registration date and a link to the SEC page are all on the row.

ABOUT THE FIRM IDENTIFIER
-------------------------
Act! has no usable firm key. Every contact carries a `companyID` field and it is
EMPTY on all 47,466 of them -- these are Contact records with a free-text company
name rather than records linked to Act! Company entities. So the Act! side offers
`act_company` and `act_department` (a branch name, where one was filed) and
nothing more solid. The SEC side supplies a real `sec_firm_crd`, which is the
identifier worth propagating back INTO Act! once a match is confirmed.
"""
from __future__ import annotations

import argparse
import pathlib
import re

import pandas as pd

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT = ROOT / "data" / "output"
INTERIM = ROOT / "data" / "interim"

# Corporate noise that says nothing about whether two firm names are the same.
NOISE = {"llc", "inc", "incorporated", "llp", "lp", "corp", "corporation", "co",
         "company", "ltd", "plc", "pc", "pa", "the", "and", "of", "group",
         "financial", "services", "service", "securities", "advisors", "advisers",
         "advisory", "wealth", "management", "capital", "partners", "associates",
         "investments", "investment", "planning", "brokerage"}


def firm_tokens(name: str) -> set:
    words = re.split(r"[^a-z0-9]+", str(name or "").lower())
    return {w for w in words if w and w not in NOISE and len(w) > 1}


def firm_agreement(act_company: str, sec_firm: str) -> str:
    """A reviewer's first question is "same firm?" -- answered here so they do not
    have to squint at two strings. A heuristic, and labelled as one."""
    a, b = firm_tokens(act_company), firm_tokens(sec_firm)
    if not a or not b:
        return "unknown"
    if a == b or a <= b or b <= a:
        return "yes"
    return "partial" if a & b else "no"



# Common short forms. Not exhaustive and not meant to be -- it exists so that
# "Bob Smith" against "ROBERT SMITH" is not presented to a human as a mystery.
NICKNAMES = {
    "bob": "robert", "rob": "robert", "bobby": "robert", "bill": "william",
    "will": "william", "billy": "william", "jim": "james", "jimmy": "james",
    "mike": "michael", "dave": "david", "dan": "daniel", "danny": "daniel",
    "tom": "thomas", "tommy": "thomas", "steve": "stephen", "chris": "christopher",
    "chuck": "charles", "charlie": "charles", "rick": "richard", "dick": "richard",
    "rich": "richard", "ted": "edward", "ed": "edward", "eddie": "edward",
    "tony": "anthony", "joe": "joseph", "joey": "joseph", "ken": "kenneth",
    "greg": "gregory", "jeff": "jeffrey", "matt": "matthew", "nick": "nicholas",
    "pat": "patrick", "sam": "samuel", "andy": "andrew", "drew": "andrew",
    "ben": "benjamin", "alex": "alexander", "tim": "timothy", "ron": "ronald",
    "don": "donald", "doug": "douglas", "larry": "lawrence", "gerry": "gerald",
    "jerry": "gerald", "fred": "frederick", "art": "arthur", "phil": "philip",
    "ray": "raymond", "walt": "walter", "vic": "victor", "cathy": "catherine",
    "kathy": "katherine", "sue": "susan", "liz": "elizabeth", "beth": "elizabeth",
    "peggy": "margaret", "meg": "margaret", "maggie": "margaret", "jen": "jennifer",
    "jenny": "jennifer", "becky": "rebecca", "debbie": "deborah", "deb": "deborah",
}


def canon(word: str) -> str:
    w = re.sub(r"[^a-z]", "", str(word or "").lower())
    return NICKNAMES.get(w, w)


def first_name_agreement(act_name: str, sec_full: str, sec_legal: str) -> str:
    """Does the Act! first name correspond to either SEC name form?

    Nearly every row in this tier was demoted for a first-name disagreement, and
    that single label covers two completely different situations: "Ronald Grove"
    against "RONALD Grove JR", which is the same man with a suffix, and "Casey
    Compston" against "MARK COMPSTON", which is not. Separating them is the
    difference between a worksheet and a wall of text.
    """
    a = str(act_name or "").strip().split()
    if not a:
        return "unknown"
    first = canon(a[0])
    if not first:
        return "unknown"
    # ONLY the leading token of each SEC name form.
    #
    # This previously also took the second token, reasoning that somebody might
    # go by their middle name. But sec_full_name already resolves that -- it is
    # built from used_first_name -- so all the second token added was the middle
    # INITIAL, and comparing a first name against a middle initial matches on one
    # shared letter. It promoted "Henry Millian" against "PATRICK H MILLIAN" to
    # an accept, which is a different person.
    candidates = set()
    for name in (sec_full, sec_legal):
        parts = str(name or "").strip().split()
        if parts:
            c = canon(parts[0])
            if c:
                candidates.add(c)
    if not candidates:
        return "unknown"
    if first in candidates:
        return "yes"
    # An initial standing in for a full name, in either direction. Weak on its
    # own, which is why it never reaches ACCEPT without the firm agreeing too.
    if len(first) == 1 and any(c.startswith(first) for c in candidates):
        return "initial"
    if any(len(c) == 1 and first.startswith(c) for c in candidates):
        return "initial"
    return "no"



NOTES = [
    ("", ""),
    ("What this is", "Every Act! contact, paired with the SEC record the automated pass "
                     "proposed for them, plus the evidence behind it."),
    ("What to do", "Fill in the verdict column: yes, no, or unsure. An accepted row gives you "
                   "candidate_crd for the person and sec_firm_crd for their firm."),
    ("tier", "high = the automated pass was confident, and these are worth spot-checking "
             "rather than reworking. review = a candidate that needs a second opinion. "
             "none = nothing found, so this needs a search rather than a decision."),
    ("suggestion", "ACCEPT, REJECT or CHECK, derived only from whether the firm and the first "
                   "name agree. A way to sort the work, not an answer. Rows are grouped by it "
                   "within each tier."),
    ("EIC assets", "act_eic_assets comes from Act! custom field user10 and is populated for "
                   "3,755 contacts. Sorting on it puts the people you already do business with "
                   "at the top, which is where the matching effort is worth most."),
    ("Postal code beats firm name", "postals_agree is the most decisive column here. A firm "
                                    "name disagrees whenever an adviser trades under their own "
                                    "brand while filing through a larger RIA -- Cosmo Boyd is "
                                    "'The 680 Group' in Act! and 'Independent Advisor Alliance' "
                                    "to the SEC -- but the Atlanta postal code matches either way."),
    ("Read the gap, not just the score", "match_gap is how far ahead this candidate was of the "
                                         "runner-up. A score of 1.0 with a gap of 0.02 means two "
                                         "people fit almost equally well -- a coin toss wearing a "
                                         "confident number."),
    ("Namesakes", "How many people in the SEC feed share the name. With 15 namesakes the name "
                  "proves nothing and the firm or postal code has to carry the decision."),
    ("Two name forms", "sec_full_name is the conversational name, sec_legal_name the filed one. "
                       "They differ whenever somebody goes by their middle name, which is common "
                       "-- Montague Laffitte Boyd III is Cosmo Boyd to everyone who knows him. "
                       "Do not reject a row on the first name alone."),
    ("Branch vs head office", "sec_office_* is the branch the adviser works from. sec_firm_hq_* "
                              "is the firm's registered address, which for anyone affiliated "
                              "with a large RIA is a head office in another state -- the two "
                              "disagree for about 70% of advisers. Compare the Act! city against "
                              "the BRANCH."),
    ("Firm families", "firms_agree compares words, not ownership. 'Bank of America Private Bank' "
                      "against 'Merrill Lynch' reads as 'no' even though one owns the other, and "
                      "the same goes for any firm since acquired or renamed. Do not batch-reject "
                      "on that column alone."),
    ("Why so many first names disagree", "One label hides two different things: 'Ronald Grove' "
                                         "against 'RONALD Grove JR' is the same man, 'Casey "
                                         "Compston' against 'MARK COMPSTON' is not. "
                                         "first_names_agree separates them."),
    ("Act! firm key", "There is none. companyID is empty on all 47,466 contacts, so the Act! "
                      "side offers only a free-text company name. sec_firm_crd is the real firm "
                      "identifier and is worth writing back into Act!."),
]


def same_postal(a, b) -> str:
    """Five-digit comparison. Act! holds ZIP+4 in places and the SEC does not, so
    comparing the full strings would report a disagreement that is not one."""
    a5 = re.sub(r"[^0-9]", "", str(a or ""))[:5]
    b5 = re.sub(r"[^0-9]", "", str(b or ""))[:5]
    if len(a5) < 5 or len(b5) < 5:
        return "unknown"
    return "yes" if a5 == b5 else "no"


def same_city(a, b) -> str:
    ca = re.sub(r"[^a-z]", "", str(a or "").lower())
    cb = re.sub(r"[^a-z]", "", str(b or "").lower())
    if not ca or not cb:
        return "unknown"
    # Act! is hand-typed and misspellings are common -- "Sheman" for "Sherman"
    # appears in the export. A prefix match on either side catches those without
    # pretending to do fuzzy matching properly.
    return "yes" if ca == cb or ca.startswith(cb) or cb.startswith(ca) else "no"


COLUMNS = [
    ("act_id", "Act! contact id. Paste into Act! to open the record."),
    ("tier", "How the automated pass rated its own match. high = confident, review = a "
             "candidate worth a second opinion, none = nothing found."),
    ("act_name", "Name as held in Act!."),
    ("act_company", "Company as held in Act!. Free text -- Act! has no company key on these records."),
    ("act_department", "Branch or department in Act!, where one was filed."),
    ("act_street", "Business street address in Act!."),
    ("act_city", "Business city in Act!."),
    ("act_state", "Business state in Act!."),
    ("act_postal", "Business postal code in Act!."),
    ("act_email", "Email in Act!. Included only to identify the record, not as SEC data."),
    ("act_eic_assets", "Total EIC assets for this contact, from Act! custom field user10. "
                       "Blank for the great majority. Sort on it to do the rows that matter most."),
    ("act_mail_code", "Act! Mail Code. U means they asked to unsubscribe, N unreachable or "
                      "bouncing, NC no mail by request."),
    ("verdict", "BLANK for you to fill in: yes / no / unsure."),
    ("firms_agree", "Heuristic: do the Act! and SEC firm names share meaningful words? yes / partial / no."),
    ("first_names_agree", "Heuristic: does the Act! first name match either SEC name form, "
                          "allowing for common short forms? yes / initial / no."),
    ("postals_agree", "Do the Act! and SEC BRANCH postal codes match on their first five "
                      "digits? The single most decisive column on the sheet."),
    ("cities_agree", "Do the Act! and SEC branch cities match, tolerating the odd typo?"),
    ("suggestion", "Read from the two columns above only. ACCEPT where both agree, REJECT where "
                   "both disagree, CHECK otherwise. A prompt, not a decision."),
    ("match_score", "How strongly the automated pass believed this match, 0 to 1."),
    ("match_gap", "Distance to the runner-up candidate. A LOW gap is the real warning: "
                  "0.02 means a second person scored almost as well."),
    ("namesakes", "How many people in the SEC feed share this name. 15 namesakes means a "
                  "name match alone is worthless."),
    ("demoted_because", "Why the automated pass would not call it confident."),
    ("candidate_crd", "The proposed SEC CRD. This is what gets written back to Act! if you accept it."),
    ("sec_full_name", "Preferred name on the SEC record."),
    ("sec_legal_name", "Legal name on the SEC record. Differs when someone goes by a middle name."),
    ("sec_firm_crd", "SEC CRD of the firm. The identifier worth carrying back into Act!."),
    ("sec_firm_name", "Firm name on the SEC record."),
    ("sec_office_street", "Street of the BRANCH the adviser works from."),
    ("sec_office_city", "Branch city -- where they sit."),
    ("sec_office_state", "Branch state -- where they sit."),
    ("sec_office_postal", "Branch postal code. Often the fastest confirmation."),
    ("sec_branch_count", "How many offices this adviser is attached to. Where there is more "
                          "than one, the office columns show whichever matches the Act! record."),
    ("sec_firm_hq_city", "City of the firm's REGISTERED address, which is frequently a head "
                         "office elsewhere. Do not compare this against the Act! city."),
    ("sec_firm_hq_state", "State of the firm's registered address."),
    ("sec_registered_since", "Earliest registration date at this firm."),
    ("sec_designations", "CFP, CFA and similar."),
    ("sec_years_experience", "Years since first exam."),
    ("iapd_url", "The SEC page. Opening it settles almost every doubtful row."),
]


def sec_side() -> pd.DataFrame:
    """One row per individual: their most recent employment, since that is the
    one a current CRM record is most likely to be about."""
    import export_individuals as ei

    df = ei.build()
    df = df.sort_values("registered_since", na_position="first")
    df = df.drop_duplicates("advisor_crd", keep="last")
    return df.rename(columns={
        "advisor_crd": "candidate_crd",
        "full_name": "sec_full_name", "legal_name": "sec_legal_name",
        "firm_crd": "sec_firm_crd", "firm_name": "sec_firm_name",
        "office_street": "sec_office_street", "office_city": "sec_office_city",
        "office_state": "sec_office_state", "office_postal": "sec_office_postal",
        "firm_hq_city": "sec_firm_hq_city", "firm_hq_state": "sec_firm_hq_state",
        "registered_since": "sec_registered_since", "designations": "sec_designations",
        "years_experience": "sec_years_experience",
    })


def build(tier: str = "") -> pd.DataFrame:
    """Every Act! contact by default.

    This started as the review tier alone, which was the wrong shape for the job:
    somebody assigning a CRD to EVERY contact needs the confident matches in
    front of them too, if only to spot-check and accept them in bulk. Montague
    "Cosmo" Boyd matched at high confidence and was therefore missing from a
    workbook whose purpose was to leave nobody unassigned.
    """
    x = pd.read_parquet(INTERIM / "act_crosswalk.parquet")
    if tier:
        x = x[x["tier"] == tier]
    x = x.copy()
    x["advisor_crd"] = x["advisor_crd"].astype(str).str.strip()

    sec = sec_side()
    sec["candidate_crd"] = sec["candidate_crd"].astype(str).str.strip()

    x["advisor_crd"] = x["advisor_crd"].replace("", pd.NA)
    df = x.rename(columns={
        "name": "act_name", "email": "act_email", "company": "act_company",
        "state": "act_state", "demoted": "demoted_because", "advisor_crd": "candidate_crd",
    }).merge(sec, on="candidate_crd", how="left")

    extras = act_extras()
    for field in ("act_department", "act_street", "act_city", "act_postal",
                  "act_eic_assets", "act_mail_code"):
        df[field] = df["act_id"].map(lambda i: (extras.get(str(i)) or {}).get(field))

    df["firms_agree"] = [firm_agreement(a, b) for a, b
                         in zip(df.get("act_company", ""), df.get("sec_firm_name", ""))]
    df["first_names_agree"] = [first_name_agreement(a, f, l) for a, f, l
                               in zip(df["act_name"], df["sec_full_name"], df["sec_legal_name"])]

    def suggest2(firms, names, postal, city):
        """Location now carries most of the weight it deserves.

        A matching postal code is the strongest single piece of evidence
        available -- far stronger than a firm name, which disagrees whenever an
        adviser trades under their own brand. It also rescues the many rows where
        Act! records the practice and the SEC records the affiliating RIA.
        """
        if names == "no" and postal != "yes":
            return "REJECT"
        if names in ("yes", "initial") and (postal == "yes" or firms == "yes" or city == "yes"):
            return "ACCEPT"
        return "CHECK"

    # Re-point the SEC office at whichever of this person's branches matches the
    # CRM record, before any comparison is made.
    branches = branch_map()
    chosen, counts = [], []
    for crd, ap, ac, st, ci, stt, po in zip(
            df["candidate_crd"], df["act_postal"], df["act_city"],
            df["sec_office_street"], df["sec_office_city"],
            df["sec_office_state"], df["sec_office_postal"]):
        cur = {"street": st, "city": ci, "state": stt, "postal": po}
        pick, n = best_branch(branches.get(str(crd)), ap, ac, cur)
        chosen.append(pick)
        counts.append(n)
    df["sec_office_street"] = [c["street"] for c in chosen]
    df["sec_office_city"] = [c["city"] for c in chosen]
    df["sec_office_state"] = [c["state"] for c in chosen]
    df["sec_office_postal"] = [c["postal"] for c in chosen]
    df["sec_branch_count"] = counts

    df["postals_agree"] = [same_postal(a, b) for a, b
                           in zip(df["act_postal"], df["sec_office_postal"])]
    df["cities_agree"] = [same_city(a, b) for a, b
                          in zip(df["act_city"], df["sec_office_city"])]

    df["suggestion"] = [suggest2(f, n, p, c) for f, n, p, c
                        in zip(df["firms_agree"], df["first_names_agree"],
                               df["postals_agree"], df["cities_agree"])]
    df["verdict"] = ""

    # CRDs are identifiers, not quantities. The tier "none" rows have no
    # candidate, and their blanks upcast the column to float -- so every CRD in
    # the sheet arrived as 845934.0, which is wrong on sight and would be worse
    # once somebody pasted it back into Act!.
    for c in ("candidate_crd", "sec_firm_crd"):
        df[c] = (df[c].astype("object")
                 .map(lambda v: "" if pd.isna(v) else str(v).strip())
                 .str.replace(r"\.0$", "", regex=True))

    for c, _ in COLUMNS:
        if c not in df.columns:
            df[c] = ""
    df = df[[c for c, _ in COLUMNS]]

    # Grouped by suggestion so the two easy piles can be worked in bulk and the
    # genuinely uncertain rows are not scattered among them. match_score is a poor
    # sort key here: it saturates at 1.0 for thousands of rows, including obvious
    # rejects, so ordering by it alone put "Casey Compston / MARK COMPSTON" at the
    # very top of the list.
    # Confident matches first, so the bulk-acceptable work is done before anyone
    # reaches the rows that need thinking about. Within a tier, grouped by the
    # triage suggestion for the same reason.
    tier_order = {"high": 0, "review": 1, "none": 2}
    order = {"ACCEPT": 0, "REJECT": 1, "CHECK": 2}
    df["_t"] = df["tier"].map(tier_order).fillna(3)
    df["_o"] = df["suggestion"].map(order).fillna(3)
    # Within a tier and triage group, biggest EIC relationship first. Getting a
    # CRD wrong on a $25M relationship costs more than on a cold prospect.
    df["_a"] = pd.to_numeric(df["act_eic_assets"], errors="coerce").fillna(-1)
    df = df.sort_values(["_t", "_o", "_a", "namesakes", "match_gap"],
                        ascending=[True, True, False, True, False])
    df = df.drop(columns=["_a"])
    return df.drop(columns=["_t", "_o"])



def branch_map() -> dict:
    """advisor_crd -> every branch on file for them.

    An adviser at a wirehouse is commonly attached to two: the firm's flagship
    office and the one they actually work from. advisor_employments names only
    one of them, and for Merrill advisers it names One Bryant Park -- so Matt
    Hepola, who sits at 80 S 8th St in Minneapolis exactly as Act! records,
    appeared to be in New York and his postal code appeared to disagree.

    Carrying all of them lets the comparison below pick the branch that actually
    corresponds to the CRM record, which is the honest question: does ANY office
    this person works from match what we hold?
    """
    path = OUT / "advisor_branches.parquet"
    if not path.exists():
        return {}
    b = pd.read_parquet(path)
    b["advisor_crd"] = b["advisor_crd"].astype(str)
    out = {}
    for row in b.itertuples(index=False):
        out.setdefault(row.advisor_crd, []).append({
            "street": getattr(row, "branch_street1", "") or "",
            "city": getattr(row, "branch_city", "") or "",
            "state": getattr(row, "branch_state", "") or "",
            "postal": getattr(row, "branch_postal", "") or "",
        })
    return out


def best_branch(branches, act_postal, act_city, current):
    """The branch that corresponds to the CRM record, if one does.

    Preference order is postal, then city, then whatever was already there. A
    postal match is close to proof; a city match is good; neither means we should
    not start substituting addresses on a hunch.
    """
    if not branches:
        return current, len(branches or [])
    for b in branches:
        if same_postal(b["postal"], act_postal) == "yes":
            return b, len(branches)
    for b in branches:
        if same_city(b["city"], act_city) == "yes":
            return b, len(branches)
    return current, len(branches)


def act_extras() -> dict:
    """act_id -> the fields the crosswalk does not carry, from the raw export.

    Address and EIC assets both live here rather than in the crosswalk, and both
    matter: an address is the most decisive evidence available for confirming a
    match, and the asset figure decides which rows are worth doing first.
    """
    import sys
    sys.path.insert(0, str(ROOT / "src"))
    from build_act_mail_codes import objects

    src = sorted((ROOT / "data" / "raw").glob("act_contacts_*.json"))
    if not src:
        return {}
    out = {}
    with src[-1].open(encoding="utf-8") as fh:
        for rec in objects(fh):
            addr = rec.get("businessAddress") or {}
            cf = rec.get("customFields") or {}
            street = " ".join(str(addr.get(k) or "").strip()
                              for k in ("line1", "line2") if addr.get(k)).strip()
            # user10 holds total EIC assets as a dollar string. Verified against
            # the application's own asset file: a contact reading 4773096.04 is
            # the advisor whose acv 4,772,343 and mf 753.04 sum to exactly that.
            raw = str(cf.get("user10") or "").replace(",", "").replace("$", "").strip()
            try:
                assets = float(raw) if raw else None
            except ValueError:
                assets = None
            out[str(rec.get("id") or "")] = {
                "act_department": str(rec.get("department") or "").strip(),
                "act_street": street,
                "act_city": str(addr.get("city") or "").strip(),
                "act_postal": str(addr.get("postalCode") or "").strip(),
                "act_eic_assets": assets,
                "act_mail_code": str(cf.get("email__y_n") or "").strip().upper(),
            }
    return out


def write_xlsx(df, path):
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header = [c for c, _ in COLUMNS]
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Contacts")
    ws.append(header)
    for row in df.itertuples(index=False, name=None):
        ws.append(["" if pd.isna(v) else v for v in row])

    doc = wb.create_sheet("How to use this")
    doc.append(["Column", "Meaning"])
    for name, meaning in COLUMNS:
        doc.append([name, meaning])
    for note in NOTES:
        doc.append(list(note))
    wb.save(path)

    wb2 = load_workbook(path)
    fill = PatternFill("solid", fgColor="FFF3CD")
    for name in ("Contacts", "How to use this"):
        s = wb2[name]
        for cell in s[1]:
            cell.font = Font(bold=True)
        s.freeze_panes = "A2"
    s = wb2["Contacts"]
    widths = {"act_id": 38, "act_name": 24, "act_company": 32, "act_department": 20,
              "act_email": 30, "verdict": 10, "tier": 9, "firms_agree": 12,
              "first_names_agree": 17, "suggestion": 12, "demoted_because": 40,
              "sec_full_name": 24, "sec_legal_name": 26, "sec_firm_name": 32,
              "sec_office_street": 30, "iapd_url": 54}
    for i, name in enumerate(header, start=1):
        s.column_dimensions[get_column_letter(i)].width = widths.get(name, 14)
    v = header.index("verdict") + 1
    for r in range(2, s.max_row + 1):
        s.cell(row=r, column=v).fill = fill
    s.auto_filter.ref = "A1:%s1" % get_column_letter(len(header))
    d = wb2["How to use this"]
    d.column_dimensions["A"].width = 30
    d.column_dimensions["B"].width = 104
    for row in d.iter_rows(min_row=2, max_col=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    wb2.save(path)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.split(chr(10))[0])
    ap.add_argument("--out", default=str(OUT / "crm_crd_matching.xlsx"))
    ap.add_argument("--tier", default="", help="limit to one tier: high, review or none")
    args = ap.parse_args()

    df = build(args.tier)
    path = pathlib.Path(args.out)
    path.parent.mkdir(parents=True, exist_ok=True)
    write_xlsx(df, path)
    print("[*] %s Act! contacts -> %s (%.1f MB)"
          % (format(len(df), ","), path, path.stat().st_size / 1e6))
    print("[*] by tier:")
    for k, v in df["tier"].value_counts().items():
        print("      %-8s %s" % (k, format(int(v), ",")))
    print("[*] triage suggestion:")
    for k, v in df["suggestion"].value_counts().items():
        print("      %-8s %s" % (k, format(int(v), ",")))



if __name__ == "__main__":
    main()
