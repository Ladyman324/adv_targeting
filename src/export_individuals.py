"""Export SEC-registered individuals to Excel, for matching against the CRM.

PURPOSE
-------
Act! holds people without CRD numbers. This file is the other side of that
match: every individual in the SEC's IAPD adviser feed, with enough surrounding
detail to disambiguate common names by hand or by script.

Deliberately WIDER than strictly necessary. A match on "Michael Brown" is
worthless on its own; a match on "Michael Brown, Raymond James, Atlanta GA
30326, registered since 2011" is decisive. The extra columns cost nothing to
carry and are the difference between a confident match and a coin toss.

NOT INCLUDED
------------
Email addresses -- those come from a separate vendor feed rather than the SEC,
and were asked for separately. Phone numbers likewise: the SEC feed has none.

GRAIN
-----
One row per advisor PER FIRM. Somebody registered at two firms appears twice,
because both employments are real and either might be the one a CRM record
refers to. Deduplicate on advisor_crd for one row per person.
"""
from __future__ import annotations

import argparse
import pathlib

import pandas as pd

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT = ROOT / "data" / "output"

# Column -> what it means, shown to whoever does the matching. This order is the
# order of the sheet: identity, then firm, then place, then context.
DICTIONARY = [
    ("advisor_crd", "SEC CRD number for the individual. This is the value you are trying to assign."),
    ("full_name", "The name a person would write: preferred first + last (+ suffix). "
                  "Edison Tate Lambeth appears here as Tate Lambeth."),
    ("legal_name", "First + middle + last (+ suffix) exactly as filed. Match against BOTH "
                   "name columns -- a CRM record may hold either form."),
    ("first_name", "Legal first name as filed."),
    ("used_first_name", "The name they actually go by, where it differs from the legal one."),
    ("middle_name", "Middle name or initial as filed."),
    ("last_name", "Surname as filed."),
    ("suffix", "Jr, III, and so on."),
    ("firm_crd", "CRD of the firm they are registered with."),
    ("firm_name", "Firm name exactly as it appears on the individual's record."),
    ("office_street", "Street of the BRANCH this person works from. Where they sit."),
    ("office_city", "Branch city. Where they sit."),
    ("office_state", "Branch state. Where they sit."),
    ("office_postal", "Branch postal code. Where they sit -- usually the fastest confirmation."),
    ("firm_hq_city", "City of the firm's REGISTERED address. Often a head office in another "
                     "state entirely; it disagrees with the branch for about 70% of advisers."),
    ("firm_hq_state", "State of the firm's registered address."),
    ("firm_hq_postal", "Postal code of the firm's registered address."),
    ("n_branch_locations", "How many branch locations this person is attached to."),
    ("registration_status", "Registration status on the adviser record."),
    ("registered_since", "Earliest registration date for this employment. Useful for ruling out a namesake."),
    ("registration_categories", "What they are registered as."),
    ("n_state_registrations", "How many states they hold an adviser registration in."),
    ("registered_states", "Which states, pipe-separated. Where they are LICENSED, not where they sit."),
    ("designations", "CFP, CFA, ChFC and similar, as filed."),
    ("years_experience", "Years since their first qualifying exam."),
    ("first_exam_date", "Date of that first exam."),
    ("n_prior_firms", "How many firms they have been registered with previously."),
    ("n_exams", "How many exams they have passed."),
    ("active_ag_reg", "Whether they hold an active agent registration."),
    ("iapd_url", "Public SEC page for this individual. The fastest way to settle a doubtful match."),
]

NOTES = [
    ("Source", "SEC IAPD investment adviser representative feed."),
    ("Grain", "One row per individual PER FIRM. Someone registered at two firms "
              "appears twice; deduplicate on advisor_crd for one row each."),
    ("Not included", "Email addresses (separate vendor file) and phone numbers "
                     "(not present in the SEC feed)."),
    ("Careful", "office_state is where they SIT. registered_states is where they are "
                "LICENSED. The two differ for roughly a fifth of advisers, so matching "
                "on the wrong one will produce confident nonsense."),
    ("Matching tip", "Try advisor_crd first if you have one, then last_name + firm_crd, "
                     "then last_name + office_postal. Fall back to a name alone only with a "
                     "second field agreeing, because common names are genuinely common."),
    ("Two name columns", "Match against full_name AND legal_name. Many advisers go by their "
                         "middle name -- Edison Tate Lambeth is 'Tate Lambeth' to everyone "
                         "who knows him -- so a CRM built from conversation holds one form "
                         "and a CRM built from paperwork holds the other."),
]


def build() -> pd.DataFrame:
    adv = pd.read_parquet(OUT / "advisors.parquet")
    emp = pd.read_parquet(OUT / "advisor_employments.parquet")

    adv["advisor_crd"] = adv["advisor_crd"].astype(str)
    emp["advisor_crd"] = emp["advisor_crd"].astype(str)
    emp["firm_crd"] = emp["firm_crd"].astype(str)

    # LEFT from advisors: somebody with no current employment row is still a
    # person the CRM might hold, and dropping them silently shrinks the pool
    # available to match against.
    df = adv.merge(emp, on="advisor_crd", how="left")

    # The BRANCH address -- where the adviser actually sits.
    #
    # advisor_employments carries the firm's REGISTERED address in emp_*, which
    # for anyone affiliated with a large RIA is a head office in another state:
    # Montague "Cosmo" Boyd files through Independent Advisor Alliance in
    # Charlotte NC while working from 1735 Peachtree St NE in Atlanta. The two
    # disagree for 69.5% of rows, so labelling emp_* as the office address would
    # be wrong far more often than right.
    #
    # Joined on the branch city and state the employment record itself names, so
    # an adviser attached to several branches gets the one belonging to this
    # employment rather than an arbitrary first pick.
    br_path = OUT / "advisor_branches.parquet"
    if br_path.exists():
        br = pd.read_parquet(br_path)
        br["advisor_crd"] = br["advisor_crd"].astype(str)
        br["firm_crd"] = br["firm_crd"].astype(str)
        br = br.drop_duplicates(["advisor_crd", "firm_crd", "branch_city", "branch_state"])
        br = br[["advisor_crd", "firm_crd", "branch_city", "branch_state",
                 "branch_street1", "branch_postal"]]
        df = df.merge(br, on=["advisor_crd", "firm_crd", "branch_city", "branch_state"], how="left")
    else:
        df["branch_street1"] = pd.NA
        df["branch_postal"] = pd.NA

    exp_path = OUT / "advisor_experience.parquet"
    if exp_path.exists():
        ex = pd.read_parquet(exp_path)[["advisor_crd", "years_experience", "first_exam_date"]]
        ex["advisor_crd"] = ex["advisor_crd"].astype(str)
        df = df.merge(ex, on="advisor_crd", how="left")
    else:
        df["years_experience"] = pd.NA
        df["first_exam_date"] = pd.NA

    used = df["used_first_name"].fillna("").astype(str).str.strip()
    first = df["first_name"].fillna("").astype(str).str.strip()
    middle = df["middle_name"].fillna("").astype(str).str.strip()
    last = df["last_name"].fillna("").astype(str).str.strip()
    suffix = df["suffix"].fillna("").astype(str).str.strip()

    # TWO name forms, because a CRM record may hold either and matching on one
    # alone throws away half the matches.
    #
    # full_name deliberately omits the middle name. A great many advisers go by
    # their middle name -- Edison TATE Lambeth, John PAUL Power -- and for those
    # people used_first_name and middle_name are the same word, so including
    # both produced "TATE TATE LAMBETH". It also is not how anyone writes their
    # own name.
    tidy = lambda s: s.str.replace(r"\s+", " ", regex=True).str.strip()
    df["full_name"] = tidy(used.where(used != "", first) + " " + last + " " + suffix)
    df["legal_name"] = tidy(first + " " + middle + " " + last + " " + suffix)

    df = df.rename(columns={
        "firm_name_on_record": "firm_name",
        "emp_city": "firm_hq_city",
        "emp_state": "firm_hq_state",
        "emp_postal": "firm_hq_postal",
        "branch_city": "office_city",
        "branch_state": "office_state",
        "branch_street1": "office_street",
        "branch_postal": "office_postal",
        "reg_status": "registration_status",
        "reg_earliest_date": "registered_since",
        "reg_categories": "registration_categories",
        "n_registrations": "n_state_registrations",
        "reg_states": "registered_states",
    })

    cols = [c for c, _ in DICTIONARY]
    for c in cols:
        if c not in df.columns:
            df[c] = pd.NA
    df = df[cols]

    # Sorted by surname, so a person scrolling for a name lands near it.
    return df.sort_values(["last_name", "first_name", "advisor_crd"], na_position="last")


def write_xlsx(df: pd.DataFrame, path: pathlib.Path) -> None:
    """Streamed with write_only: 420,000 rows through openpyxl's normal path is
    a memory problem rather than merely a slow one."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    header = [c for c, _ in DICTIONARY]
    wb = Workbook(write_only=True)

    ws = wb.create_sheet("Individuals")
    ws.freeze_panes = "A2"
    ws.append(header)
    for row in df.itertuples(index=False, name=None):
        ws.append(["" if pd.isna(v) else v for v in row])

    doc = wb.create_sheet("What each column is")
    doc.append(["Column", "Meaning"])
    for name, meaning in DICTIONARY:
        doc.append([name, meaning])
    doc.append([])
    for name, meaning in NOTES:
        doc.append([name, meaning])
    wb.save(path)

    # Header styling and column widths need a second pass -- a write_only sheet
    # cannot style rows as it streams them.
    wb2 = load_workbook(path)
    sheet = wb2["Individuals"]
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for i, name in enumerate(header, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = min(max(len(name) + 3, 12), 34)
    sheet.auto_filter.ref = "A1:%s1" % get_column_letter(len(header))
    doc2 = wb2["What each column is"]
    for cell in doc2[1]:
        cell.font = Font(bold=True)
    doc2.column_dimensions["A"].width = 26
    doc2.column_dimensions["B"].width = 108
    wb2.save(path)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("--out", default=str(OUT / "sec_individuals.xlsx"))
    ap.add_argument("--state", help="limit to one office state, e.g. GA")
    ap.add_argument("--csv", action="store_true", help="also write a .csv alongside")
    args = ap.parse_args()

    df = build()
    if args.state:
        df = df[df["office_state"].fillna("").astype(str).str.upper() == args.state.upper()]

    path = pathlib.Path(args.out)
    path.parent.mkdir(parents=True, exist_ok=True)
    write_xlsx(df, path)
    print("[*] %s rows -> %s (%.1f MB)" % (format(len(df), ","), path, path.stat().st_size / 1e6))
    print("[*] %s distinct individuals" % format(df["advisor_crd"].nunique(), ","))
    if args.csv:
        c = path.with_suffix(".csv")
        df.to_csv(c, index=False)
        print("[*] csv -> %s (%.1f MB)" % (c, c.stat().st_size / 1e6))


if __name__ == "__main__":
    main()
